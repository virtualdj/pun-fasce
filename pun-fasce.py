import argparse
import requests, zipfile, io
from openpyxl import load_workbook
from datetime import date, timedelta
import holidays
from statistics import mean

# Restituisce il numero di fascia oraria
def get_fascia(data, festivo, ora):
	#F1 = lu-ve 8-19
	#F2 = lu-ve 7-8, lu-ve 19-23, sa 7-23
	#F3 = lu-sa 0-7, lu-sa 23-24, do, festivi
	if festivo or (data.weekday() == 6):
		# Festivi e domeniche
		return 3
	elif (data.weekday() == 5):
		# Sabato
		if (ora >= 7) and (ora < 23):
			return 2
		else:
			return 3
	else:
		# Altri giorni della settimana
		if (ora == 7) or ((ora >= 19) and (ora < 23)):
			return 2
		elif (ora == 23) or ((ora >= 0) and (ora < 7)):
			return 3
	return 1

# Formatta la media come numero decimale con 6 decimali (ma arrotondato al quinto)
def fmt_mean(list):
	return format(round(mean(list), 5), '.6f')

# Calcola la fascia F23 sulla base delle fasce F2 e F3, formattando il risultato
def calc_f23(f2, f3):
	# La motivazione del calcolo è oscura, ma sembra corretta, vedere:
	# https://github.com/virtualdj/pun_sensor/issues/24#issuecomment-1829846806

	return format(round(mean(f2), 5) * .46 + round(mean(f3), 5) * .54, '.6f')

# Recupera l'anno dall'argomento (o usa quello corrente)
parser = argparse.ArgumentParser(description='Mostra i costi del PUN di un anno, diviso per mesi e fasce orarie.')
parser.add_argument('year', type=int, nargs='?',
	default = date.today().year,
	help='Anno da processare (4 cifre)')
args = parser.parse_args()

# Supportato il formato dal 2016 in poi
if (args.year < 2016) or (args.year > date.today().year):
	raise SystemExit('ERRORE: l\'anno specificato non è valido/supportato (%s).' % args.year)
else:
	anno = args.year

# Scarica il file ZIP con i dati dal sito Mercato Elettrico
try:
	req = requests.get(f'https://www.mercatoelettrico.org/it-it/Home/Esiti/Elettricita/MGP/Statistiche/DatiStorici/moduleId/10874/controller/GmeDatiStoriciItem/action/DownloadFile?fileName=Anno{anno}.zip')
except:
	raise SystemExit('ERRORE: sito web MercatoElettrico.org non disponibile!')

# Scompatta lo ZIP in memoria
try:
	archive = zipfile.ZipFile(io.BytesIO(req.content))
except:
	raise SystemExit('ERRORE: il file scaricato dal sito non è valido.')

# Verifica se ci sono i prezzi ogni 15 minuti nell'archivio
xlsFileName = ''
for f in archive.filelist:
	if f.filename.lower().endswith('_60.xlsx'):
		xlsFileName = f.filename
		break

# Se non ci sono, estrae il primo file Excel disponibile
if (xlsFileName == ''):
	for f in archive.filelist:
		if f.filename.lower().endswith('.xlsx'):
			xlsFileName = f.filename
			break

# Verifica se almeno un nome di file è stato trovato
if (xlsFileName != ''):
  
	# Recupera il file non compresso in memoria
	xlfile = archive.open(xlsFileName)
	
	# Apre il file Excel
	try:
		wb = load_workbook(xlfile)
	except:
		raise SystemExit('ERRORE: file Excel non valido nel file scaricato dal sito.')

	# Trova il foglio prezzi
	try:
		sheet = wb['Prezzi-Prices']
	except:
		raise SystemExit('ERRORE: file Excel \'Prezzi\' non trovato nel file scaricato dal sito.')

	# Carica le festività
	it_holidays = holidays.IT()

	# Inizializza le variabili di conteggio
	festivo = False
	prev_dat = ''
	prev_month = 0
	f1 = []
	f2 = []
	f3 = []
	monoorario = []

	# Header output
	print('Mese','MO (€/kWh)','F1 (€/kWh)','F2 (€/kWh)','F3 (€/kWh)','F23 (€/kWh)', sep='\t')

	# Esamina le righe non vuote a partire dalla seconda
	for row in range(2, sheet.max_row + 1):
		if (sheet.cell(row, 1).value is None):
			break

		# Estrae i valori delle celle dell'Excel
		dat = str(sheet.cell(row, 1).value) #YYYYMMDD
		ora = sheet.cell(row, 2).value - 1 # 1..24
		prezzo = sheet.cell(row, 3).value / 1000

		# Verifica se il giorno è cambiato
		if (dat != prev_dat):
			# Converte la stringa giorno in data
			dat2 = date(int(dat[0:4]), int(dat[4:6]), int(dat[6:8]))
	
			# Verifica la festività
			festivo = dat2 in it_holidays
	
			# Aggiorna il valore di data
			prev_dat = dat

		# Verifica se il mese è cambiato
		if (dat2.month != prev_month):
			# Nuovo mese	
			# Stampa i totali precedenti
			if (prev_month > 0):
				print(f'{prev_month}/{anno}', fmt_mean(monoorario), fmt_mean(f1), fmt_mean(f2), fmt_mean(f3), calc_f23(f2, f3), sep='\t')
		
			# Memorizza il nuovo mese
			prev_month = dat2.month
	
			# Azzera i conteggi
			f1.clear()
			f2.clear()
			f3.clear()
			monoorario.clear()

		# Estrae la fascia oraria
		#print("Len", len(f1), len(f2), len(f3))
		fascia = get_fascia(dat2, festivo, ora)
		if fascia == 3:
			f3.append(prezzo)
		elif fascia == 2:
			f2.append(prezzo)
		elif fascia == 1:
			f1.append(prezzo)
		monoorario.append(prezzo)

	# Verifica se l'ultimo mese è completo
	next_day = dat2 + timedelta(days=1)
	if (next_day.month != prev_month):
		# Mese completo, mostra statistiche
		print(f'{prev_month}/{anno}', fmt_mean(monoorario), fmt_mean(f1), fmt_mean(f2), fmt_mean(f3), calc_f23(f2, f3), sep='\t')

# Chiude l'archivio
archive.close()

# Mostra un errore se non è stato trovato alcun file
if (xlsFileName == ''):
	raise SystemExit('ERRORE: nessun file Excel è stato trovato nel file scaricato dal sito.')
